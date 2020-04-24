import json
from openpyxl import load_workbook
from collections import OrderedDict
from dataContextJSON import getDataContexts

wb = load_workbook("BRMv3.xlsx", data_only=True)
ws = wb["Base OC's"]


# Loop through entire sheet creating a dictionary
# for each row (element)
elements = OrderedDict()
firstRowFlag = True
for row in ws.iter_rows():
    if firstRowFlag is True:
        firstRowFlag = False
        continue

    IDref = row[0].value
    objectClass = row[1].value
    elementName = row[2].value
    
    dataContextList = row[5].value.split("~~")
    definition = row[6].value.split("~~")
    HCDRM = row[7].value.split("~~")
    Exchange = row[8].value.split("~~")
    managingPartner = row[9].value.split("~~")
    exampleValue = row[10].value.split("~~")
    opmStandardMapping = row[11].value.split("~~")
    formMapping = row[12].value.split("~~")
    owner = row[13].value.split("~~")
    sorn = row[14].value.split("~~")

    # List of data items
    items = {"Beta Extract Element Definition": definition,
             "Beta Extract Proposed HC Data Reference Model (HCDRM) Code": HCDRM,
             "Exchange Summary": Exchange,
             "Beta Extract Proposed LOB Managing Partner": managingPartner,
             "Example Value": exampleValue,
             "OPM Standards Mapping": opmStandardMapping,
             "Form Mapping": formMapping,
             "Proposed Steward": owner,
             "SORN": sorn}

    # Create json object for each element
    elements[IDref] = OrderedDict()

    for i, dataContext in enumerate(dataContextList):

        elements[IDref][dataContext] = OrderedDict()
        elements[IDref][dataContext]["Object Class"] = objectClass
        elements[IDref][dataContext]["Element Name"] = elementName
        for name, itemList in items.iteritems():
            try:
                elements[IDref][dataContext][name] = itemList[i]
            except:
                elements[IDref][dataContext][name] = "Data is missing"


# Write the JSON file
with open('elementJSON.json', 'w') as outfile:
    json.dump(elements, outfile)


# Call the script to get the data contexts from the 
# HCIM
print "Finished getting all the base elements..."
getDataContexts()
