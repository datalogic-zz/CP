import os
from lxml import etree
from openpyxl import load_workbook
from collections import OrderedDict
from utils import convertToJSON

wb = load_workbook("BRMv3.xlsx", data_only=True)
ws = wb["SubFunctions + Services + Class"]
element_ws = wb["Base OC's"]

# Generate first level of dictionary with only
# the functions.
functionList = []
BRM = OrderedDict()
for row in ws.iter_rows():
    function = row[0].value.strip()
    if function not in functionList:
        BRM[function] = OrderedDict()
        functionList.append(function)

# Link the subfunctions to the functions that are
# already in the dictionary.
for function in functionList:
    subfunctionList = []
    for row in ws.iter_rows():
        subfunction = row[1].value.strip()

        if subfunction not in subfunctionList and function == row[0].value.strip():
            BRM[function][subfunction] = OrderedDict()
            subfunctionList.append(subfunction)

# Link the services to their appropriate subfunctions now
serviceList = []
for function, subfunctions in BRM.items():
    for subfunction in subfunctions:
        for row in ws.iter_rows():
            service = row[2].value.strip()
            if row[0].value.strip() == function:
                if row[1].value.strip() == subfunction:
                    BRM[function][subfunction][service] = []

# Run through the entire sheet and add the object classes into
# their appropriate lists.
for row in ws.iter_rows():
    function = row[0].value.strip()
    subfunction = row[1].value.strip()
    service = row[2].value.strip()

    objectClassList = BRM[function][subfunction][service]
    counter = 0
    for cell in row:
        if counter < 3:
            counter += 1
        else:
            if cell.value is not None:
                objectClassList.append(cell.value)


# The BRM dictionary contains:
# function(dict) --> subfunction(dict) --> service(dict) --> object classes(dict)

# Begin writing to XML document
root = etree.Element("root")
firstName = etree.SubElement(root, "name")
firstName.text = "HCBRM"
etree.SubElement(root, "visualName").text = "HCBRM"
etree.SubElement(root, "legendName").text = "HCBRM"
uniqueID = 0
etree.SubElement(root, "uniqueID").text = str(uniqueID)

# Loop through the first level of the BRM dictionary
for function, subfunctions in sorted(BRM.items()):
    print function
    firstChildTag = etree.SubElement(root, "children")

    firstName = etree.SubElement(firstChildTag, "name")
    firstName.text = function.split(" ", 1)[1]
    etree.SubElement(firstChildTag, "visualName").text = function.split(" ", 1)[1]
    etree.SubElement(firstChildTag, "legendName").text = "HCBRM : " + function.split(" ", 1)[1]
    uniqueID += 1
    etree.SubElement(firstChildTag, "uniqueID").text = str(uniqueID)

    # Loop through the second level 
    for subfunction, services in sorted(subfunctions.items()):
        secondChildTag = etree.SubElement(firstChildTag, "children")
        subfunctionName = etree.SubElement(secondChildTag, "name")
        subfunctionName.text = subfunction.split(" ", 1)[1]
        etree.SubElement(secondChildTag, "visualName").text = subfunction.split(" ", 1)[1]
        etree.SubElement(secondChildTag, "legendName").text = "HCBRM : " + function.split(" ", 1)[1] + " : " + subfunction.split(" ", 1)[1]
        uniqueID += 1
        etree.SubElement(secondChildTag, "uniqueID").text = str(uniqueID)

        # Loop through the third level
        for service, objClasses in sorted(services.items()):
            if service != "TBD":
                thirdChildTag = etree.SubElement(secondChildTag, "children")
                serviceName = etree.SubElement(thirdChildTag, "name")
                serviceName.text = service.split(" ", 1)[1]
                etree.SubElement(thirdChildTag, "visualName").text = service.split(" ", 1)[1]
                etree.SubElement(thirdChildTag, "legendName").text = "HCBRM : " + function.split(" ", 1)[1] + " : " + subfunction.split(" ", 1)[1] + " : " + service.split(" ", 1)[1]
                flag = 1
                uniqueID += 1
                etree.SubElement(thirdChildTag, "uniqueID").text = str(uniqueID)

            else:
                thirdChildTag = etree.SubElement(secondChildTag, "children")
                serviceName = etree.SubElement(thirdChildTag, "name")
                serviceName.text = service
                etree.SubElement(thirdChildTag, "visualName").text = service
                etree.SubElement(thirdChildTag, "legendName").text = "HCBRM : " + function.split(" ", 1)[1] + " : " + subfunction.split(" ", 1)[1] + " : " + service
                flag = 2
                uniqueID += 1
                etree.SubElement(thirdChildTag, "uniqueID").text = str(uniqueID)

            # Loop through the object class list
            for obj in sorted(objClasses):
                fourthChildTag = etree.SubElement(thirdChildTag, "children")
                objName = etree.SubElement(fourthChildTag, "name")
                objName.text = obj
                etree.SubElement(fourthChildTag, "visualName").text = obj
                if flag == 1:
                    etree.SubElement(fourthChildTag, "legendName").text = "HCBRM : " + function.split(" ", 1)[1] + " : " + subfunction.split(" ", 1)[1] + " : " + service.split(" ", 1)[1] + " : " + obj
                    uniqueID += 1
                    etree.SubElement(fourthChildTag, "uniqueID").text = str(uniqueID)

                elif flag == 2:
                    etree.SubElement(fourthChildTag, "legendName").text = "HCBRM : " + function.split(" ", 1)[1] + " : " + subfunction.split(" ", 1)[1] + " : " + service + " : " + obj
                    uniqueID += 1
                    etree.SubElement(fourthChildTag, "uniqueID").text = str(uniqueID)

                # Activate sheet that contains the object classes and elements
                counter = 0
                for row in element_ws.iter_rows():
                    if obj == row[1].value:
                        counter += 1

                        fifthChildTag = etree.SubElement(fourthChildTag, "children")
                        elementName = etree.SubElement(fifthChildTag, "name")
                        elementName.text = row[0].value
                        etree.SubElement(fifthChildTag, "visualName").text = str(counter) + ": " + row[2].value
                        etree.SubElement(fifthChildTag, "size").text = str(len(str(row[2].value)))


                        etree.SubElement(fifthChildTag, "elementChecker").text = "true"
                        etree.SubElement(fifthChildTag, "idNumber").text = str(row[0].value)
                        uniqueID += 1
                        etree.SubElement(fifthChildTag, "uniqueID").text = str(uniqueID)


# Write the XML and JSON files out
tree = etree.ElementTree(root)
tree.write("BRM Functions temp.xml",
           pretty_print=True)

# Open the xml file and remove any: "  :" and ":  "
with open("BRM Functions.xml", "w") as writeFile:
    with open("BRM Functions temp.xml", "r") as readFile:
        lines = readFile.readlines()
        for line in lines:
            line = line.replace("  :", " :")
            line = line.replace(":  ", ": ")
            writeFile.write(line)

# Remove the temporary file
os.remove("BRM Functions temp.xml")

# Convert the xml doc to a json doc
convertToJSON("BRM Functions.xml")
print "Number of total elements: " + str(uniqueID+1)
