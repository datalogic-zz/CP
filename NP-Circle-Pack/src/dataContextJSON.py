## RUN elementJSONgenerator.py, NOT THIS FILE ##

import json
from openpyxl import load_workbook
import copy

def printMatches(obj, dataContext, element):
	objString = "Obj: " + obj
	dataContextString = ". Data Context: " + dataContext
	elementString = ". Element: " + element

	print objString + dataContextString + elementString

def removeDataContext(stripString, string):
	'''
	This function will strip off the data context
	from ISO Concept Label
	ie. "Hard Copy Check Address Line Text" will
	become "Address Line Text"
	'''

	if stripString is not None:
		return string.replace(stripString + " ", "")
	elif stripString is None:
		return string

def findColumnIndex(columnName, ws):
	'''
	This function is going to find the column
	index for each meta data column; if the column
	is not found by name, it will return None
	'''

	firstRow = ws[1]
	for index in firstRow:
		if index.value == columnName:
			return index.col_idx - 1

	# If the column name isn't found in the 
	# first row
	return None


def getCellValue(row, index):
	'''
	The function that will get the cell value
	given the index for that column; if the index
	is none, that means the column was not found earlier
	and will return "missing value" for that specific meta
	data.
	'''

	try:
		return row[index].value
	except:
		return "missing value"

def getDataContexts():
	print "Beginning to get the data contexts..."

	# Read JSON data into the datastore variable
	filename = "elementJSON.json"
	if filename:
	    with open(filename, 'r') as f:
	        datastore = json.load(f)

	# Read in the workbook
	wb = load_workbook("ISO Columns.xlsx")
	ws = wb["HCIM Version"]

	# Assign indices for the columns
	objIdx = findColumnIndex("ISO Object Class", ws)
	dataContextIdx = findColumnIndex("ISO Data Context", ws)
	elmIdx = findColumnIndex("ISO Data Element Concept Label", ws)

	# Assign indices for the meta data columns
	formMapIdx = findColumnIndex("Form Mapping", ws) # Holding off for now, fix in HCIM
	defIdx = findColumnIndex("Proposed Element Definition", ws)
	exchangeIdx = findColumnIndex("Exchange Summary", ws)
	ownerIdx = findColumnIndex("Proposed Steward", ws) # Also another option for accepted steward..
	sornIdx = findColumnIndex("SORN", ws) # Holding off for now, fix in HCIM to be in one column
	opmStdMapIdx = findColumnIndex("OPM Standards Mapping", ws)
	exampleIdx = findColumnIndex("Example Value", ws)
	managingPartnerIdx = findColumnIndex("Beta Extract Proposed LOB Managing Partner", ws)
	hcdrmIdx = findColumnIndex("Beta Extract Proposed HC Data Reference Model (HCDRM) Code", ws)


	# Go row by row on the sheet and add
	# the object classes
	hcimDict = {}
	for row in ws.iter_rows(min_row=2):
		if row[objIdx].value not in hcimDict:
			hcimDict[row[objIdx].value] = {}

	# Add the data contexts to the object classes
	for obj in hcimDict:
		for row in ws.iter_rows(min_row=2):

			if row[objIdx].value is not None and row[dataContextIdx].value is not None:
				if row[objIdx].value == obj:
					if row[dataContextIdx].value not in hcimDict[obj] and \
					row[dataContextIdx].value.upper() != "BASE":
						if row[dataContextIdx].value != "":
							hcimDict[obj][row[dataContextIdx].value] = {}

	# Remove all the object classes
	# that don't have a data context
	deleteList = []
	for obj in hcimDict:
		if len(hcimDict[obj]) == 0:
			deleteList.append(obj)
	for d in deleteList:
		del hcimDict[d]

	# Loop through the HCIM to get the
	# elements that fall under the obj class + data context
	for i, obj in enumerate(hcimDict):
		print str(i + 1) + " " + str(len(hcimDict))
		for dataContext in hcimDict[obj]:
			for row in ws.iter_rows(min_row=2):

				if row[dataContextIdx].value is not None and row[elmIdx].value is not None:
					element = removeDataContext(row[dataContextIdx].value, row[elmIdx].value)

					if row[objIdx].value == obj and row[dataContextIdx].value == dataContext:
						if element not in hcimDict[obj][dataContext]:
							hcimDict[obj][dataContext][element] = {}

							# Add dummy data for now
							hcimDict[obj][dataContext][element] = {}
							hcimDict[obj][dataContext][element][dataContext] = {}

							# Add all of the element data for that specific data 
							# context and element and class
							hcimDict[obj][dataContext][element][dataContext]["Object Class"] = obj
							hcimDict[obj][dataContext][element][dataContext]["Element Name"] = element
							
							# The chunk of code below is for the meta data for each
							# element
							hcimDict[obj][dataContext][element][dataContext]["Form Mapping"] = getCellValue(row, formMapIdx)
							hcimDict[obj][dataContext][element][dataContext]["Beta Extract Element Definition"] = getCellValue(row, defIdx)
							hcimDict[obj][dataContext][element][dataContext]["Exchange Summary"] = getCellValue(row, exchangeIdx)
							hcimDict[obj][dataContext][element][dataContext]["Proposed Steward"] = getCellValue(row, ownerIdx)
							hcimDict[obj][dataContext][element][dataContext]["SORN"] = getCellValue(row, sornIdx)
							hcimDict[obj][dataContext][element][dataContext]["OPM Standards Mapping"] = getCellValue(row, opmStdMapIdx)
							hcimDict[obj][dataContext][element][dataContext]["Example Value"] = getCellValue(row, exampleIdx)
							hcimDict[obj][dataContext][element][dataContext]["Beta Extract Proposed LOB Managing Partner"] = getCellValue(row, managingPartnerIdx)
							hcimDict[obj][dataContext][element][dataContext]["Beta Extract Proposed HC Data Reference Model (HCDRM) Code"] = getCellValue(row, hcdrmIdx)

	# Need to create a copy of the dictionary
	# because you cannot edit the dictionary
	# while looping through it
	newDatastore = copy.deepcopy(datastore)

	# Loop through the json file
	for ID in datastore:
		for base in datastore[ID]:
			elementName = datastore[ID][base]["Element Name"]
			objectClass = datastore[ID][base]["Object Class"]

			# Loop through the hcim dictionary
			for obj in hcimDict:
				if obj == objectClass:
					for dataContext in hcimDict[obj]:
						for element in hcimDict[obj][dataContext]:
							if element == elementName:
								printMatches(obj, dataContext, element)
								newDatastore[ID][dataContext] = hcimDict[obj][dataContext][element][dataContext]


	# print json.dumps(hcimDict, indent=4, sort_keys=True)
	with open('elementJSON.json', 'w') as outfile:
	    json.dump(newDatastore, outfile)

if __name__ == "__main__":
	# getDataContexts()
	print "Run the elementJSONgenerator file, not this one..."
